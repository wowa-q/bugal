"""
Gehört zu busines layer. spezialisierter Handler.

"""

from datetime import datetime
import logging

from openpyxl import Workbook
# from openpyxl.utils import exceptions as openpyxl_exception

# from bugal import cfg
from . import bugal_if as a
from cfg import config as cfg
from libs import exceptions as err


logger = logging.getLogger(__name__)


class ExcelWriter(a.HandlerWriteIF):
    """Handler class to print data to excel file
    """

    def __init__(self, xls_file):
        self.xls_file = xls_file
        self._work_book = Workbook()
        self._work_book.iso_dates = True
        self.transactions = []
        self.history = None
        self.rule = None
        self.properties = None
        self.guide = ""

    def print_transactions(self):
        """create transactions sheet and print transaction to the excel
        """
        # create the sheet
        sheet = self._work_book.create_sheet('Transaktionen')

        # print header
        sheet.cell(row=cfg.MIN_ROW - 1, column=cfg.COLUMNS['Datum'], value='Datum')
        sheet.cell(row=cfg.MIN_ROW - 1, column=cfg.COLUMNS['Buchungstext'], value='Buchungstext')
        # pylint: disable=C0301
        sheet.cell(row=cfg.MIN_ROW - 1, column=cfg.COLUMNS['Verwendungszweck'], value='Verwendungszweck')
        sheet.cell(row=cfg.MIN_ROW - 1, column=cfg.COLUMNS['Kontonummer'], value='Kontonummer')
        sheet.cell(row=cfg.MIN_ROW - 1, column=cfg.COLUMNS['Status'], value='Status')
        sheet.cell(row=cfg.MIN_ROW - 1, column=cfg.COLUMNS['Betrag'], value='Betrag')
        sheet.cell(row=cfg.MIN_ROW - 1, column=cfg.COLUMNS['Checksum'], value='Checksum')
        sheet.cell(row=cfg.MIN_ROW - 1, column=cfg.COLUMNS['vom Konto'], value='vom Konto')
        # check if transaction is an iterable
        try:
            iter(self.transactions)
        except TypeError:
            raise err.NoValidTransactionData

        # print data
        for t_ctr, transaction in enumerate(self.transactions, start=cfg.MIN_ROW):
            sheet.cell(row=t_ctr, column=cfg.COLUMNS['Datum'], value=transaction.datum)
            sheet.cell(row=t_ctr, column=cfg.COLUMNS['Buchungstext'], value=transaction.text)
            # pylint: disable=C0301
            sheet.cell(row=t_ctr, column=cfg.COLUMNS['Verwendungszweck'], value=transaction.verwendung)
            sheet.cell(row=t_ctr, column=cfg.COLUMNS['Kontonummer'], value=transaction.konto)
            sheet.cell(row=t_ctr, column=cfg.COLUMNS['Status'], value=transaction.status)
            sheet.cell(row=t_ctr, column=cfg.COLUMNS['Betrag'], value=transaction.value)
            sheet.cell(row=t_ctr, column=cfg.COLUMNS['Checksum'], value=hash(transaction))
            sheet.cell(row=t_ctr, column=cfg.COLUMNS['vom Konto'], value=transaction.src_konto)

    def print_history(self):
        """Print history to excel
        """
        sheet = self._work_book.create_sheet('Historie')
        sheet.cell(row=2, column=2, value='not implemented')

    def print_rules(self):
        """Print rules to excel
        """
        sheet = self._work_book.create_sheet('Regeln')
        sheet.cell(row=2, column=2, value='not implemented')

    def print_properties(self):
        """Print Properties to Excel
        """
        sheet = self._work_book.create_sheet('Properties')
        sheet.cell(row=2, column=2, value='not implemented')

    def print_year(self):
        """Prints an overview for the year to excel
        """
        sheet = self._work_book.create_sheet('Jahr')
        sheet.cell(row=2, column=2, value='not implemented')

    def print_guide(self):
        """prints the users guide to excel

        #     Returns:
        #         None
        """
        guide = """
                In der Tabelle Transactions findest du alle Transaktionen,
                die aus der Batenbank exportiert worden sind"""
        sheet = self._work_book.create_sheet('Guide')
        sheet.cell(row=2, column=2, value=guide)

    def print(self):
        return self.write_all_to_excel()

    def write_all_to_excel(self):
        """Creates sheets for all relevant data (transactions,
                                                history,
                                                rules,
                                                properties,
                                                guide,
                                                year),
        #     and prints the data to the Excel workbook.

        #     Returns:
        #         bool: True, when writing was siccessful
        """

        success = False

        try:
            for method_name in ['print_transactions',
                                'print_history',
                                'print_properties',
                                'print_guide',
                                'print_rules',
                                'print_year']:
                method = getattr(self, method_name)
                method()
            try:
                self._work_book.remove('Sheet')
            except ValueError:
                logger.exception("Could not delete sheet with the name: %s", 'Sheet')
            self._work_book.save(self.xls_file)
            self._work_book.close()
            success = True

            # except PermissionError as permission:
            #     print(f"An error occurred while writing to the Excel file: {permission}, \
            #           please close the excel file and press enter")
            #     input()

        except ValueError as value:
            logger.exception("An error occurred while writing to the Excel file: %s", value)
            self._work_book.close()
        return success

    def __repr__(self) -> str:
        return f"{self.__class__.__name__}to write into Excel located in {self.xls_file}"
