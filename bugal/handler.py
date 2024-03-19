"""Handlers
    - xls reades & writer
    - csv importer
    - artifact handler
"""

# from dataclasses import fields, asdict
import os
import hashlib
import zipfile
import csv
import pathlib
from datetime import datetime
import logging

from openpyxl import Workbook
# from openpyxl.utils import exceptions as openpyxl_exception

from bugal import cfg
from bugal import abstract as a
from bugal import exceptions as err


logger = logging.getLogger(__name__)


class ExcelWriter(a.HandlerWriteIF):
    """Handler class to print data to excel file
    """

    def __init__(self, xls_file):
        self.xls_file = xls_file
        self._work_book = Workbook()
        self._work_book.iso_dates = True
        self.transactions = []
        self.history = None
        self.rule = None
        self.properties = None
        self.guide = ""

    def print_transactions(self):
        """create transactions sheet and print transaction to the excel
        """
        # create the sheet
        sheet = self._work_book.create_sheet('Transaktionen')

        # print header
        sheet.cell(row=cfg.MIN_ROW - 1, column=cfg.COLUMNS['Datum'], value='Datum')
        sheet.cell(row=cfg.MIN_ROW - 1, column=cfg.COLUMNS['Buchungstext'], value='Buchungstext')
        # pylint: disable=C0301
        sheet.cell(row=cfg.MIN_ROW - 1, column=cfg.COLUMNS['Verwendungszweck'], value='Verwendungszweck')
        sheet.cell(row=cfg.MIN_ROW - 1, column=cfg.COLUMNS['Kontonummer'], value='Kontonummer')
        sheet.cell(row=cfg.MIN_ROW - 1, column=cfg.COLUMNS['Status'], value='Status')
        sheet.cell(row=cfg.MIN_ROW - 1, column=cfg.COLUMNS['Betrag'], value='Betrag')
        sheet.cell(row=cfg.MIN_ROW - 1, column=cfg.COLUMNS['Checksum'], value='Checksum')
        sheet.cell(row=cfg.MIN_ROW - 1, column=cfg.COLUMNS['vom Konto'], value='vom Konto')
        # print data
        for t_ctr, transaction in enumerate(self.transactions, start=cfg.MIN_ROW):
            sheet.cell(row=t_ctr, column=cfg.COLUMNS['Datum'], value=transaction.date)
            sheet.cell(row=t_ctr, column=cfg.COLUMNS['Buchungstext'], value=transaction.text)
            # pylint: disable=C0301
            sheet.cell(row=t_ctr, column=cfg.COLUMNS['Verwendungszweck'], value=transaction.verwendung)
            sheet.cell(row=t_ctr, column=cfg.COLUMNS['Kontonummer'], value=transaction.konto)
            sheet.cell(row=t_ctr, column=cfg.COLUMNS['Status'], value=transaction.status)
            sheet.cell(row=t_ctr, column=cfg.COLUMNS['Betrag'], value=transaction.value)
            sheet.cell(row=t_ctr, column=cfg.COLUMNS['Checksum'], value=hash(transaction))
            sheet.cell(row=t_ctr, column=cfg.COLUMNS['vom Konto'], value=transaction.src_konto)

    def print_history(self):
        """Print history to excel
        """
        sheet = self._work_book.create_sheet('Historie')
        sheet.cell(row=2, column=2, value='not implemented')

    def print_rules(self):
        """Print rules to excel
        """
        sheet = self._work_book.create_sheet('Regeln')
        sheet.cell(row=2, column=2, value='not implemented')

    def print_properties(self):
        """Print Properties to Excel
        """
        sheet = self._work_book.create_sheet('Properties')
        sheet.cell(row=2, column=2, value='not implemented')

    def print_year(self):
        """Prints an overview for the year to excel
        """
        sheet = self._work_book.create_sheet('Jahr')
        sheet.cell(row=2, column=2, value='not implemented')

    def print_guide(self):
        """prints the users guide to excel

        #     Returns:
        #         None
        """
        guide = """
                In der Tabelle Transactions findest du alle Transaktionen,
                die aus der Batenbank exportiert worden sind"""
        sheet = self._work_book.create_sheet('Guide')
        sheet.cell(row=2, column=2, value=guide)

    def print(self):
        return self.write_all_to_excel()

    def write_all_to_excel(self):
        """Creates sheets for all relevant data (transactions,
                                                history,
                                                rules,
                                                properties,
                                                guide,
                                                year),
        #     and prints the data to the Excel workbook.

        #     Returns:
        #         bool: True, when writing was siccessful
        """

        success = False

        try:
            for method_name in ['print_transactions',
                                'print_history',
                                'print_properties',
                                'print_guide',
                                'print_rules',
                                'print_year']:
                method = getattr(self, method_name)
                method()
            try:
                self._work_book.remove('Sheet')
            except ValueError:
                logger.exception("Could not delete sheet with the name: %s", 'Sheet')
            self._work_book.save(self.xls_file)
            self._work_book.close()
            success = True

            # except PermissionError as permission:
            #     print(f"An error occurred while writing to the Excel file: {permission}, \
            #           please close the excel file and press enter")
            #     input()

        except ValueError as value:
            logger.exception("An error occurred while writing to the Excel file: %s", value)
            self._work_book.close()
        return success

    def __repr__(self) -> str:
        return f"{self.__class__.__name__}to write into Excel located in {self.xls_file}"


class CSVImporter(a.HandlerReadIF):
    """Handler class to import data csv file(s)
    """
    csv_hashes = []

    def __init__(self, pth, input_type=None):
        self.pth = pathlib.Path(pth)
        self.csv_files = []
        if self.pth.is_file():
            self.csv_files.append(pth)
        else:
            self.csv_files = list(self.pth.glob('**/*.csv'))
        self.meta_data = []
        # input type setting
        if cfg.TYPE == 'BETA':
            self.input_type = cfg.TransactionListBeta
        elif cfg.TYPE == 'CLASSIC':
            self.input_type = cfg.TransactionListClassic
        else:
            self.input_type = input_type
        logger.info("CSV Handler initialized")

    def get_meta_data(self):
        if os.path.isfile(self.pth):
            self.meta_data.append(self._get_meta_data(self.pth))
        else:
            # TODO: to be tested
            files = list(self.pth.glob('**/*.csv'))
            for fl_pth in files:
                self.meta_data.append(self._get_meta_data(fl_pth))
        return self.meta_data

    def _get_meta_data(self, csv_file):
        """returns some meta data necessary for creation a history Dataclass instance

        Returns:
            list: list of dictionpories. Every dict should have:
            - 'end_date': datetime object in format '%d.%m.%Y'
            - 'start_date': datetime object in format '%d.%m.%Y'
            - 'account': string object showing from which account the data were exported
            - 'checksum': calculated checksum over the csv file to be compared with checksums from History table in DB
            - 'file_name': (str) showing the file name
            - 'file_ext': showing the file extension
        """
        if self.input_type is None:
            raise err.NoInputTypeSet('Meta data collection: Input type not configured')
        meta = cfg.CSV_META.copy()
        meta['checksum'] = self.get_checksum(csv_file)
        meta['file_ext'] = 'csv'
        meta['file_name'] = str(csv_file).strip('.csv')
        meta['start_date'] = datetime.strptime('01.01.3000', "%d.%m.%Y")
        meta['end_date'] = datetime.strptime('01.01.1000', "%d.%m.%Y")
        for ctr, line in enumerate(self.read_lines(csv_file)):
            if len(line) < 2:
                continue
            if ctr == 0:
                meta['account'] = line[1].replace("Girokonto", "").replace("/", "").strip()
            elif ctr > self.input_type.CSV_START_ROW.value:
                date_str = line[0]
                date_obj = self._get_date_object(date_str)
                if date_obj:
                    if date_obj < meta['start_date']:
                        meta['start_date'] = date_obj
                    if date_obj > meta['end_date']:
                        meta['end_date'] = date_obj
            else:
                logger.warning("Kein toter Code, line: %s", ctr)
        return meta

    def _get_date_object(self, date_str: str) -> datetime:
        try:
            if len(date_str) == 8:
                date_obj = datetime.strptime(date_str, "%d.%m.%y")
            elif len(date_str) == 10:
                date_obj = datetime.strptime(date_str, "%d.%m.%Y")
            else:
                return False
            return date_obj
        except ValueError as error:
            logger.debug("Date could not be extracted from CSV line: %s", date_str)
            logger.exception("A %s has occurred", type(error).__name__)
            raise ValueError(f"datum provided with false format: {date_str}") from error
        except KeyError as error:
            logger.debug("Dictionary does not have the key")
            logger.exception("A %s has occurred", type(error).__name__)
        except AttributeError as error:
            logger.exception("A %s has occurred", type(error).__name__)
            logger.debug("start_date or end_date is not datetime")
        except TypeError as error:
            logger.exception("A %s has occurred", type(error).__name__)
            logger.debug("start_date or end_date is not datetime")

    def read_lines(self, csv_file):
        with open(csv_file, encoding='ISO-8859-1') as csvfile:
            reader = csv.reader(csvfile, delimiter=';')
            logger.info("reader initialized with encoding: %s", 'ISO-8859-1')
            for row in reader:
                yield row

    def get_checksum(self, csv_file=None):
        """provides hash value for csv file

        Args:
            csv_file (Path): csv file reference

        Returns:
            str: calculated hash value as String
        """
        if csv_file is None:
            csv_file = self.pth

        with open(csv_file, encoding='ISO-8859-1', mode='r') as csvfile:
            checksum = hashlib.md5(csvfile.read().encode('ISO-8859-1')).hexdigest().upper()
            logger.info("csv hash calculated: %s", checksum)
            return checksum

    def get_transactions(self):
        """reads line from CSV file and yields a line as transaction
            If a directory is provided, loops over csv files in the directory
        Raises:
            NoValidInputFilesFound: if no csv files were found in the given directory

        Yields:
            generator: to get the line as a list two for loops are necessary
        """
        if self.input_type is None:
            logger.debug("No input type set")
            raise err.NoInputTypeSet('CSV Import: Input type not configured')

        def read_lines(csv_file):
            with open(csv_file, mode='r', encoding='ISO-8859-15', newline='') as csvfile:
                reader = csv.reader(csvfile, delimiter=';')
                for ctr, line in enumerate(reader):
                    if ctr > self.input_type.CSV_START_ROW.value:
                        yield line

        for csv_file in self.csv_files:
            yield read_lines(csv_file)

    def archive(self):
        archiver = ArtifactHandler(cfg.ARCHIVE)
        archiver.archive_imports(self.pth)

    def __repr__(self) -> str:
        return f"{self.__class__.__name__} for the type {self.input_type} located in {self.pth}"


class ArtifactHandler(a.Artifact):
    """Handles artifacts
    """
    def __init__(self, archive=None):
        if archive is not None and archive.is_file():
            self.archive = archive
        else:
            self.archive = cfg.ARCHIVE

    def archive_imports(self, artifact=None):
        """adds the imported csv file into archive
        """
        if artifact is not None:
            with zipfile.ZipFile(self.archive, 'a', compression=zipfile.ZIP_DEFLATED) as newzip:
                newzip.write(artifact, arcname=artifact.name)
            artifact.unlink()
            logger.info("CSV file archived in %s: ", self.archive)
            return True
        else:
            return False

    def __repr__(self) -> str:
        return f"{self.__class__.__name__} for archiving in {self.archive}"
