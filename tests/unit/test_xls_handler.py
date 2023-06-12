# pylint: skip-file
# flake8: noqa

from datetime import date
import pathlib

import pytest
from openpyxl import load_workbook
import openpyxl

from context import bugal

from bugal import cfg
from bugal import handler


FIXTURE_DIR = pathlib.Path(__file__).parent.parent.resolve() / "fixtures"

#@pytest.mark.skip()
def test_excel_file_created(fx_stack_example, fx_xls_file):
    xls_writer = handler.ExcelWriter(fx_xls_file)
    # stack = model.Stack()
    # for line in fx_transactions_list_example:
    #     stack.create_transaction(line)
    xls_writer.transactions = fx_stack_example.transactions
    xls_writer.history = None
    xls_writer.rule = None
    xls_writer.properties = None
    xls_writer.guide = ""

    xls_writer.print()

    assert len(list(FIXTURE_DIR.glob('*.xlsx'))) > 0

#@pytest.mark.skip()
def test_required_sheets_created(fx_stack_example, fx_xls_file, fx_mandatory_sheets):
    xls_writer = handler.ExcelWriter(fx_xls_file)
    xls_writer.transactions = fx_stack_example.transactions
    xls_writer.history = None
    xls_writer.rule = None
    xls_writer.properties = None
    xls_writer.guide = ""

    xls_writer.print()
    work_book = load_workbook(fx_xls_file, read_only=True, keep_vba=True)
    for name in fx_mandatory_sheets:
        assert name in work_book.sheetnames, f"No {name} sheet in woork book"
    
    work_book.close()

#@pytest.mark.skip()
def test_transactions_header_printed(fx_stack_example, fx_xls_file):
    xls_writer = handler.ExcelWriter(fx_xls_file)
    xls_writer.transactions = fx_stack_example.transactions
    xls_writer.print()
    work_book = load_workbook(fx_xls_file)    
    sheet = work_book['Transaktionen']
    

    row_it = sheet.iter_rows(min_row=cfg.MIN_ROW-1, 
                               min_col=cfg.MIN_COL, 
                               max_col=30, 
                               max_row=cfg.MIN_ROW)    
    split_row_list = ([cell.value if cell.value is not None else "" for cell in row] for row in row_it)
    test_dict = {}
    for row in split_row_list:        
        for ctr, header in enumerate(cfg.COLUMNS.keys()):
            assert header in row, f"header {header} not found" 
        break

# @pytest.mark.skip()
def test_transactions_printed_in_right_column(fx_stack_example, fx_xls_file):
    xls_writer = handler.ExcelWriter(fx_xls_file)
    xls_writer.transactions = fx_stack_example.transactions
    success = xls_writer.print()
    assert success == True, "Writing to excel was not done"
    work_book = load_workbook(fx_xls_file)    
    sheet = work_book['Transaktionen']
    headers = cfg.COLUMNS.keys()
    for t_ctr, transaction in enumerate(fx_stack_example.transactions):
        for header in headers:
            val = sheet.cell(row=cfg.MIN_ROW+t_ctr, column=cfg.COLUMNS[header]).value
            if header == 'Datum':
                assert transaction.date == val, f"read value: {val}, line {cfg.MIN_ROW+t_ctr}"
            elif header == 'Buchungstext':
                assert transaction.text == val, f"read value: {val}, line {cfg.MIN_ROW+t_ctr}"
            elif header == 'Verwendungszweck':
                assert transaction.verwendung == val, f"read value: {val}, line {cfg.MIN_ROW+t_ctr}"
            elif header == 'Kontonummer':
                assert transaction.konto == val, f"read value: {val}, line {cfg.MIN_ROW+t_ctr}"
            elif header == 'BLZ':
                assert transaction.blz == val, f"read value: {val}, line {cfg.MIN_ROW+t_ctr}"
            elif header == 'Betrag':
                assert transaction.value == val, f"read value: {val}, line {cfg.MIN_ROW+t_ctr}"
            elif header == 'vom Konto':
                assert transaction.src_konto == val, f"read value: {val}, line {cfg.MIN_ROW+t_ctr}"
     
    work_book.close()
    
#@pytest.mark.skip()
def test_transactions_unique_hash_printed(fx_stack_example, fx_xls_file):
    xls_writer = handler.ExcelWriter(fx_xls_file)
    xls_writer.transactions = fx_stack_example.transactions
    success = xls_writer.print()
    assert success == True, "Writing to excel was not done"

    work_book = load_workbook(fx_xls_file)
    sheet = work_book['Transaktionen']
    column_letter = openpyxl.utils.get_column_letter(cfg.COLUMNS['Checksum'])
    column = sheet[column_letter]
    cell_values = []
    for cell in column:
        if cell.value is not None:
            cell_values.append(cell.value)
    assert len(cell_values) == len(set(cell_values)), f"duplicate vlaues exist {cell_values}"

@pytest.mark.skip()
def test_write_all_to_excel_exceptions():
        # ACHTUNG: die Dateien werden bei Intialisieren der Klasse erstellt -> Exceptions sind an der falschen stelle
        # Test the method with an invalid input file
        filename = 'nonexistent_file.xlsx'
        xls_writer = handler.ExcelWriter(None)
        with pytest.raises(FileNotFoundError):
            xls_writer.write_all_to_excel()

        # Test the method with a file that is already open
        filename = 'test_input.xlsx'
        xls_writer = handler.ExcelWriter(filename)
        with open(filename, 'rb') as file:
            with pytest.raises(PermissionError):
                xls_writer.write_all_to_excel()